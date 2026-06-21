from __future__ import annotations

import json
import random
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl

from .llm_client import LLMResponse, call_openrouter, sha256_text
from .prompt_templates import (
    PROMPT_STAGE1_REVIEW,
    format_responses_block,
    optional_context_line,
    render_prompt,
)
from .stage1_discovery import QuestionContext, load_question_and_responses
from .themes import generate_theme_id, normalize_label


def load_existing_themes(
    workbook_path: str, question_id: str
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Load existing active/candidate theme_groups and themes for the question.
    
    Returns: (theme_groups, themes) as lists of dicts with full row data.
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    
    # Load theme groups
    groups: List[Dict[str, Any]] = []
    if "theme_groups" in wb.sheetnames:
        gsheet = wb["theme_groups"]
        gheaders = [str(c).strip() for c in next(gsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        
        for row in gsheet.iter_rows(min_row=2, values_only=True):
            gdict = {gheaders[i]: row[i] for i in range(min(len(gheaders), len(row)))}
            gqid = str(gdict.get("question_id", "")).strip()
            status = str(gdict.get("status", "")).strip().lower()
            
            if gqid != question_id.strip():
                continue
            if status not in ("active", "candidate"):
                continue
            
            groups.append(gdict)
    
    # Load theme groups for mapping
    group_map: Dict[str, Dict[str, str]] = {}
    if "theme_groups" in wb.sheetnames:
        gsheet = wb["theme_groups"]
        gheaders = [str(c).strip() for c in next(gsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        for row in gsheet.iter_rows(min_row=2, values_only=True):
            gdict = {gheaders[i]: row[i] for i in range(min(len(gheaders), len(row)))}
            gid = str(gdict.get("theme_group_id", "")).strip()
            glabel = str(gdict.get("theme_group_label", "")).strip()
            gqid = str(gdict.get("question_id", "")).strip()
            if gid and gqid == question_id.strip():
                group_map[gid] = {
                    "theme_group_id": gid,
                    "theme_group_label": glabel,
                    "theme_group_definition": str(gdict.get("theme_group_definition", "")).strip(),
                }
    
    # Load themes
    themes: List[Dict[str, Any]] = []
    if "theme_catalog" in wb.sheetnames:
        tsheet = wb["theme_catalog"]
        theaders = [str(c).strip() for c in next(tsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        
        for row in tsheet.iter_rows(min_row=2, values_only=True):
            tdict = {theaders[i]: row[i] for i in range(min(len(theaders), len(row)))}
            tqid = str(tdict.get("question_id", "")).strip()
            status = str(tdict.get("status", "")).strip().lower()
            
            if tqid != question_id.strip():
                continue
            if status not in ("active", "candidate"):
                continue
            
            # Add group label if available
            gid = str(tdict.get("theme_group_id", "")).strip()
            if gid and gid in group_map:
                tdict["theme_group_label"] = group_map[gid]["theme_group_label"]
            
            themes.append(tdict)
    
    wb.close()
    return groups, themes


def _format_existing_groups_block(existing_groups: List[Dict[str, Any]]) -> str:
    lines: list[str] = []
    for g in existing_groups:
        gid = str(g.get("theme_group_id", "")).strip()
        glabel = str(g.get("theme_group_label", "")).strip()
        gdef = str(g.get("theme_group_definition", "")).strip()
        lines.append(f"- [{gid}] {glabel}: {gdef}")
    return "\n".join(lines) if lines else "(none)"


def _format_existing_themes_block(existing_themes: List[Dict[str, Any]]) -> str:
    lines: list[str] = []
    for t in existing_themes:
        tid = str(t.get("theme_id", "")).strip()
        tlabel = str(t.get("theme_label", "")).strip()
        tdef = str(t.get("theme_definition", "")).strip()
        glabel = str(t.get("theme_group_label", "")).strip() or "Ungrouped"
        lines.append(f"- [{tid}] {tlabel} (Group: {glabel}): {tdef}")
    return "\n".join(lines) if lines else "(none)"


def build_review_prompt(
    qctx: QuestionContext,
    existing_groups: List[Dict[str, Any]],
    existing_themes: List[Dict[str, Any]],
    responses: List[Dict[str, Any]],
    max_new_themes: int = 30,
) -> str:
    """Build prompt for theme review that includes existing themes and asks for validation/recommendations."""
    return render_prompt(
        PROMPT_STAGE1_REVIEW,
        survey_name=qctx.survey_name,
        question_code=qctx.question_code,
        question_text=qctx.question_text,
        question_context_section=optional_context_line("Question context", qctx.question_context),
        research_context_section=optional_context_line("Research context", qctx.research_context),
        custom_context_section=optional_context_line("Custom context", qctx.custom_context),
        responses_block=format_responses_block(responses),
        existing_groups_block=_format_existing_groups_block(existing_groups),
        existing_themes_block=_format_existing_themes_block(existing_themes),
        max_new_themes=max_new_themes,
    )


def parse_review_json(
    text: str, existing_theme_ids: set[str]
) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], List[str], List[Dict[str, str]]]:
    """Parse LLM JSON response for theme review.
    
    Returns: (validated_themes, new_themes, retirement_suggestions, replacement_themes)
    - validated_themes: List of {theme_id, action} dicts
    - new_themes: List of new theme dicts (with optional replaces_theme_id)
    - retirement_suggestions: List of theme_ids to retire
    - replacement_themes: List of new themes that replace old ones (replaces_theme_id populated)
    """
    # Try to extract JSON from markdown code blocks if present
    json_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
    if json_match:
        text = json_match.group(1)
    
    # Try to parse as-is first
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        # If parsing fails, try to find a valid JSON object starting from the first {
        start_idx = text.find('{')
        if start_idx >= 0:
            # Try to find a closing brace that creates valid JSON
            for end_idx in range(len(text), start_idx, -1):
                try:
                    candidate = text[start_idx:end_idx] + '}'
                    data = json.loads(candidate)
                    break
                except json.JSONDecodeError:
                    continue
            else:
                raise ValueError("Could not parse LLM response as JSON")
        else:
            raise ValueError("Could not find JSON object in LLM response")
    
    validated = data.get("validated_themes") or []
    new_themes_raw = data.get("new_themes") or []
    retirements = data.get("suggested_retirements") or []
    
    # Normalize validated themes
    validated_themes: List[Dict[str, str]] = []
    for v in validated:
        tid = str(v.get("theme_id", "")).strip()
        action = str(v.get("action", "")).strip().lower()
        if tid and action in ("keep", "retire"):
            validated_themes.append({"theme_id": tid, "action": action})
    
    # Normalize new themes
    new_themes: List[Dict[str, str]] = []
    replacement_themes: List[Dict[str, str]] = []
    
    for t in new_themes_raw:
        tlabel = str(t.get("theme_label", "")).strip()
        tdef = str(t.get("theme_definition", "")).strip()
        glabel = str(t.get("group_label", "")).strip()
        replaces_id = str(t.get("replaces_theme_id", "")).strip() if t.get("replaces_theme_id") else None
        
        if not tlabel:
            continue
        
        theme_dict = {
            "theme_label": tlabel,
            "theme_definition": tdef,
            "group_label": glabel,
        }
        
        if replaces_id and replaces_id in existing_theme_ids:
            theme_dict["replaces_theme_id"] = replaces_id
            replacement_themes.append(theme_dict)
        else:
            new_themes.append(theme_dict)
    
    # Normalize retirement suggestions
    retirement_suggestions: List[str] = []
    for rid in retirements:
        rid_str = str(rid).strip()
        if rid_str in existing_theme_ids:
            retirement_suggestions.append(rid_str)
    
    # Also extract retirements from validated_themes with action="retire"
    for v in validated_themes:
        if v.get("action") == "retire" and v.get("theme_id") not in retirement_suggestions:
            retirement_suggestions.append(v["theme_id"])
    
    return validated_themes, new_themes, retirement_suggestions, replacement_themes


def write_review_results(
    *,
    input_workbook: str,
    output_workbook: str,
    qctx: QuestionContext,
    validated_themes: List[Dict[str, str]],
    new_theme_groups: List[Dict[str, str]],
    new_themes: List[Dict[str, str]],
    replacement_themes: List[Dict[str, str]],
    retirement_suggestions: List[str],
    model_run_meta: Dict[str, Any],
) -> None:
    """Write review results to workbook.
    
    - New themes get status="candidate-add"
    - Themes marked for retirement get status="candidate-retire" (in-place update, preserving all data)
    - Replacement themes get status="candidate-add" with replaced_prior_theme_ids populated
    """
    src_wb = openpyxl.load_workbook(input_workbook)
    
    # Load existing data for lookups
    existing_group_label_to_id: Dict[str, str] = {}
    existing_theme_id_to_row: Dict[str, int] = {}
    
    # Load theme_groups sheet
    if "theme_groups" in src_wb.sheetnames:
        ws_g = src_wb["theme_groups"]
        headers_g = [str(c).strip() for c in next(ws_g.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        col_map_g = {h: i + 1 for i, h in enumerate(headers_g)}
        
        for row_idx, row in enumerate(ws_g.iter_rows(min_row=2, values_only=True), start=2):
            if row and len(row) >= len(headers_g):
                row_dict = {headers_g[i]: row[i] if i < len(row) else None for i in range(len(headers_g))}
                gid = str(row_dict.get("theme_group_id", "")).strip()
                glabel = str(row_dict.get("theme_group_label", "")).strip()
                gqid = str(row_dict.get("question_id", "")).strip()
                if gid and gqid == qctx.question_code.strip():
                    existing_group_label_to_id[glabel] = gid
    
    # Load theme_catalog sheet
    if "theme_catalog" in src_wb.sheetnames:
        ws_t = src_wb["theme_catalog"]
        headers_t = [str(c).strip() for c in next(ws_t.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        
        for row_idx, row in enumerate(ws_t.iter_rows(min_row=2, values_only=True), start=2):
            if row and len(row) >= len(headers_t):
                row_dict = {headers_t[i]: row[i] if i < len(row) else None for i in range(len(headers_t))}
                tid = str(row_dict.get("theme_id", "")).strip()
                tqid = str(row_dict.get("question_id", "")).strip()
                if tid and tqid == qctx.question_code.strip():
                    existing_theme_id_to_row[tid] = row_idx
    
    # Prepare metadata
    current_year = datetime.now().year
    model_name = model_run_meta.get("model", "unknown")
    current_timestamp = datetime.now()
    
    # Process retirement suggestions: update status to candidate-retire
    # Note: We update status only (not labels/definitions) to preserve longitudinal data integrity
    if retirement_suggestions and "theme_catalog" in src_wb.sheetnames:
        ws_t = src_wb["theme_catalog"]
        headers_t = [str(c).strip() for c in next(ws_t.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        col_map_t = {h: i + 1 for i, h in enumerate(headers_t)}
        status_col = col_map_t.get("status", 7)
        
        for theme_id in retirement_suggestions:
            if theme_id in existing_theme_id_to_row:
                row_num = existing_theme_id_to_row[theme_id]
                ws_t.cell(row=row_num, column=status_col, value="candidate-retire")
    
    # Find next row for new groups
    if "theme_groups" not in src_wb.sheetnames:
        ws_g = src_wb.create_sheet("theme_groups")
        headers_g = [
            "theme_group_id", "theme_group_label", "theme_group_definition",
            "question_id", "seen_years", "status", "last_edited_by", "last_edited_at", "notes"
        ]
        ws_g.append(headers_g)
        next_group_row = 2
    else:
        ws_g = src_wb["theme_groups"]
        headers_g = [str(c).strip() for c in next(ws_g.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        col_map_g = {h: i + 1 for i, h in enumerate(headers_g)}
        next_group_row = 2  # Start at row 2 (after header)
        for row_idx, row in enumerate(ws_g.iter_rows(min_row=2, values_only=True), start=2):
            if row and any(cell is not None and str(cell).strip() for cell in (row[0:3] if len(row) >= 3 else row)):
                next_group_row = row_idx + 1  # Next row after this non-empty row
            else:
                break
    
    # Process new theme groups (candidate-add)
    col_map_g = {h: i + 1 for i, h in enumerate(headers_g)}
    next_group_seq = 1
    
    for g in new_theme_groups:
        glabel = g.get("group_label", "").strip()
        gdef = g.get("group_definition", "").strip()
        if not glabel:
            continue
        if glabel in existing_group_label_to_id:
            continue  # Skip if group already exists
        
        gid = generate_theme_id(qctx.survey_name, qctx.question_code, glabel, next_group_seq)
        next_group_seq += 1
        existing_group_label_to_id[glabel] = gid
        
        ws_g.cell(row=next_group_row, column=col_map_g.get("theme_group_id", 1), value=gid)
        ws_g.cell(row=next_group_row, column=col_map_g.get("theme_group_label", 2), value=glabel)
        ws_g.cell(row=next_group_row, column=col_map_g.get("theme_group_definition", 3), value=gdef)
        ws_g.cell(row=next_group_row, column=col_map_g.get("question_id", 4), value=qctx.question_code)
        ws_g.cell(row=next_group_row, column=col_map_g.get("seen_years", 5), value=current_year)
        ws_g.cell(row=next_group_row, column=col_map_g.get("status", 6), value="candidate-add")
        ws_g.cell(row=next_group_row, column=col_map_g.get("last_edited_by", 7), value=model_name)
        ws_g.cell(row=next_group_row, column=col_map_g.get("last_edited_at", 8), value=current_timestamp)
        next_group_row += 1
    
    # Find next row for new themes
    if "theme_catalog" not in src_wb.sheetnames:
        ws_t = src_wb.create_sheet("theme_catalog")
        headers_t = [
            "theme_id", "theme_group_id", "theme_label", "theme_definition",
            "question_id", "seen_years", "status", "last_edited_by", "last_edited_at", "replaced_prior_theme_ids"
        ]
        ws_t.append(headers_t)
        next_theme_row = 2
    else:
        ws_t = src_wb["theme_catalog"]
        headers_t = [str(c).strip() for c in next(ws_t.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        next_theme_row = 2  # Start at row 2 (after header)
        for row_idx, row in enumerate(ws_t.iter_rows(min_row=2, values_only=True), start=2):
            if row and any(cell is not None and str(cell).strip() for cell in (row[0:3] if len(row) >= 3 else row)):
                next_theme_row = row_idx + 1  # Next row after this non-empty row
            else:
                break
    
    col_map_t = {h: i + 1 for i, h in enumerate(headers_t)}
    next_theme_seq = 1
    
    # Process new themes (candidate-add)
    for t in new_themes:
        tlabel = t.get("theme_label", "").strip()
        tdef = t.get("theme_definition", "").strip()
        glabel = t.get("group_label", "").strip()
        gid = existing_group_label_to_id.get(glabel) if glabel else None
        
        if not tlabel:
            continue
        
        tid = generate_theme_id(qctx.survey_name, qctx.question_code, tlabel, next_theme_seq)
        next_theme_seq += 1
        
        ws_t.cell(row=next_theme_row, column=col_map_t.get("theme_id", 1), value=tid)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("theme_group_id", 2), value=gid)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("theme_label", 3), value=tlabel)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("theme_definition", 4), value=tdef)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("question_id", 5), value=qctx.question_code)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("seen_years", 6), value=current_year)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("status", 7), value="candidate-add")
        ws_t.cell(row=next_theme_row, column=col_map_t.get("last_edited_by", 8), value=model_name)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("last_edited_at", 9), value=current_timestamp)
        # replaced_prior_theme_ids column (10) - leave empty for new themes
        next_theme_row += 1
    
    # Process replacement themes (candidate-add with replaced_prior_theme_ids)
    for t in replacement_themes:
        tlabel = t.get("theme_label", "").strip()
        tdef = t.get("theme_definition", "").strip()
        glabel = t.get("group_label", "").strip()
        replaces_id = t.get("replaces_theme_id", "").strip()
        gid = existing_group_label_to_id.get(glabel) if glabel else None
        
        if not tlabel or not replaces_id:
            continue
        
        tid = generate_theme_id(qctx.survey_name, qctx.question_code, tlabel, next_theme_seq)
        next_theme_seq += 1
        
        ws_t.cell(row=next_theme_row, column=col_map_t.get("theme_id", 1), value=tid)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("theme_group_id", 2), value=gid)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("theme_label", 3), value=tlabel)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("theme_definition", 4), value=tdef)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("question_id", 5), value=qctx.question_code)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("seen_years", 6), value=current_year)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("status", 7), value="candidate-add")
        ws_t.cell(row=next_theme_row, column=col_map_t.get("last_edited_by", 8), value=model_name)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("last_edited_at", 9), value=current_timestamp)
        ws_t.cell(row=next_theme_row, column=col_map_t.get("replaced_prior_theme_ids", 10), value=replaces_id)
        next_theme_row += 1
    
    # Write model_runs audit entry
    if "model_runs" not in src_wb.sheetnames:
        ws_m = src_wb.create_sheet("model_runs")
        ws_m.append(["model_run_id", "created_ts", "stage", "model", "temperature", "sample_size", "prompt_hash", "notes"])
    else:
        ws_m = src_wb["model_runs"]
    
    headers_m = [str(c).strip() for c in next(ws_m.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
    next_m_row = ws_m.max_row + 1
    
    col_map_m = {h: i + 1 for i, h in enumerate(headers_m)}
    ws_m.cell(row=next_m_row, column=col_map_m.get("model_run_id", 1), value=model_run_meta.get("model_run_id", ""))
    ws_m.cell(row=next_m_row, column=col_map_m.get("created_ts", 2), value=model_run_meta.get("created_ts", 0))
    ws_m.cell(row=next_m_row, column=col_map_m.get("stage", 3), value="theme_review")
    ws_m.cell(row=next_m_row, column=col_map_m.get("model", 4), value=model_run_meta.get("model", ""))
    ws_m.cell(row=next_m_row, column=col_map_m.get("temperature", 5), value=model_run_meta.get("temperature", 0.2))
    ws_m.cell(row=next_m_row, column=col_map_m.get("sample_size", 6), value=model_run_meta.get("sample_size", 0))
    ws_m.cell(row=next_m_row, column=col_map_m.get("prompt_hash", 7), value=model_run_meta.get("prompt_hash", ""))
    
    # Save
    out_path = Path(output_workbook)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    src_wb.save(str(out_path))
    src_wb.close()


def review_themes(
    workbook: str,
    *,
    question_id: str,
    model: str,
    sample_size: int | None = None,
    temperature: float = 0.2,
    max_tokens: int = 32000,
    max_new_themes: int = 30,
    output: str | None = None,
    verbose: int = 0,
    context_column: str | None = None,
) -> str:
    """Run theme review on a workbook and write results to a new file.
    
    Reviews existing active/candidate themes against the corpus, suggesting:
    - New themes to add (candidate-add)
    - Themes to retire (candidate-retire)
    - Replacement themes for modifications (candidate-add with replaced_prior_theme_ids)
    
    Returns the output workbook path.
    """
    start_time = time.time()
    
    # Load question and responses
    qctx, responses, total_responses = load_question_and_responses(
        workbook, question_id=question_id, sample_size=sample_size, context_column=context_column
    )
    
    # Load existing themes - must exist for review
    existing_groups, existing_themes = load_existing_themes(workbook, question_id)
    if not existing_themes:
        raise ValueError(
            f"No active/candidate themes found for question_id '{question_id}'. "
            "Use 'qrla discover' first to create themes, or ensure themes exist with status 'active' or 'candidate'."
        )
    
    existing_theme_ids = {str(t.get("theme_id", "")).strip() for t in existing_themes if t.get("theme_id")}
    
    if verbose >= 1:
        print(
            f"Loaded {len(existing_groups)} theme groups and {len(existing_themes)} existing themes",
            file=sys.stderr,
        )
        print(f"Sampling {len(responses)} responses from {total_responses} total...", file=sys.stderr)
        print(f"Calling {model} with review prompt...", file=sys.stderr)
    
    # Build and call LLM
    prompt = build_review_prompt(qctx, existing_groups, existing_themes, responses, max_new_themes=max_new_themes)
    
    if verbose >= 2:
        print("\n--- Review Prompt ---", file=sys.stderr)
        print(prompt, file=sys.stderr)
        print("--- End Prompt ---\n", file=sys.stderr)
    
    prompt_hash = sha256_text(prompt)
    llm_call_start = time.time()
    
    try:
        llm_resp: LLMResponse = call_openrouter(
            model=model,
            prompt=prompt,
            temperature=temperature,
            max_tokens=max_tokens,
        )
    except Exception as e:
        if verbose >= 1:
            print(f"ERROR: {e}", file=sys.stderr)
        raise
    
    llm_call_duration = time.time() - llm_call_start
    finish_reason = llm_resp.raw.get("choices", [{}])[0].get("finish_reason", "unknown")
    
    if not llm_resp.text:
        raise ValueError(
            f"LLM returned empty response. Finish reason: {finish_reason}. "
            f"This may indicate the response was truncated or the model format isn't supported."
        )
    
    if finish_reason == "length":
        print(
            f"Warning: LLM response was truncated (hit max_tokens limit). Results may be incomplete.",
            file=sys.stderr,
        )
    
    if verbose >= 2:
        print("\n--- LLM Response ---", file=sys.stderr)
        print(llm_resp.text, file=sys.stderr)
        print("--- End Response ---\n", file=sys.stderr)
    
    # Parse response
    validated_themes, new_themes, retirement_suggestions, replacement_themes = parse_review_json(
        llm_resp.text, existing_theme_ids
    )
    
    # Validate retirement suggestions
    invalid_retirements = [rid for rid in retirement_suggestions if rid not in existing_theme_ids]
    if invalid_retirements and verbose >= 1:
        print(
            f"Warning: Invalid theme_ids in retirement suggestions (ignored): {invalid_retirements}",
            file=sys.stderr,
        )
    retirement_suggestions = [rid for rid in retirement_suggestions if rid in existing_theme_ids]
    
    # Prepare new theme groups from new themes
    new_theme_groups: List[Dict[str, str]] = []
    seen_group_labels = set()
    for t in new_themes + replacement_themes:
        glabel = t.get("group_label", "").strip()
        if glabel and glabel not in seen_group_labels:
            # Check if group already exists in existing_groups
            existing_labels = {str(g.get("theme_group_label", "")).strip() for g in existing_groups}
            if glabel not in existing_labels:
                new_theme_groups.append({"group_label": glabel, "group_definition": ""})
                seen_group_labels.add(glabel)
    
    created_ts = time.time()
    model_run_id = f"{normalize_label(qctx.survey_name)}_{normalize_label(qctx.question_code)}_{int(created_ts)}"
    
    # Extract token usage
    input_tokens = 0
    output_tokens = 0
    usage = llm_resp.raw.get("usage") or {}
    if usage:
        input_tokens = usage.get("prompt_tokens", 0)
        output_tokens = usage.get("completion_tokens", 0)
    
    model_run_meta = {
        "model_run_id": model_run_id,
        "created_ts": created_ts,
        "model": llm_resp.model,
        "temperature": temperature,
        "sample_size": sample_size if sample_size is not None else 0,
        "prompt_hash": prompt_hash,
    }
    
    # Determine output path
    in_path = Path(workbook)
    if output is None:
        output = str(in_path.with_suffix("").as_posix() + f".reviewed.{normalize_label(llm_resp.model)}.xlsx")
    
    # Write results
    write_review_results(
        input_workbook=workbook,
        output_workbook=output,
        qctx=qctx,
        validated_themes=validated_themes,
        new_theme_groups=new_theme_groups,
        new_themes=new_themes,
        replacement_themes=replacement_themes,
        retirement_suggestions=retirement_suggestions,
        model_run_meta=model_run_meta,
    )
    
    total_duration = time.time() - start_time
    
    if verbose >= 1:
        print(
            f"Review complete: {len(validated_themes)} validated, {len(new_themes)} new, {len(replacement_themes)} replacements, "
            f"{len(retirement_suggestions)} retirements in {total_duration:.1f} seconds "
            f"({input_tokens} input + {output_tokens} output tokens)",
            file=sys.stderr,
        )
    
    return output

