from typing import Dict, List, Tuple
import openpyxl


def load_template_schema(template_path: str) -> Dict[str, List[str]]:
    """Load sheet -> header list mapping from the template workbook."""
    wb = openpyxl.load_workbook(template_path, read_only=True)
    schema: Dict[str, List[str]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        headers: List[str] = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c).strip() for c in row if c is not None]
            break
        schema[name] = headers
    return schema


def validate_workbook(path: str, template_path: str) -> Tuple[bool, List[str]]:
    """Validate that workbook at `path` contains required sheets and headers from template.

    Returns (is_valid, list_of_issues)
    """
    issues: List[str] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception as e:  # noqa: BLE001 - surface load failures as validation issues
        return False, [f"unable to open workbook: {e}"]

    template_schema = load_template_schema(template_path)

    # check sheets
    for sheet in template_schema:
        if sheet not in wb.sheetnames:
            issues.append(f"missing sheet: {sheet}")

    # check headers for sheets that exist
    for sheet, headers in template_schema.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        first_row: List[str] = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            first_row = [str(c).strip() for c in row if c is not None]
            break
        if headers:
            missing = [h for h in headers if h not in first_row]
            if missing:
                issues.append(f"sheet {sheet} missing headers: {missing}")

    return (len(issues) == 0), issues


