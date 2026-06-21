from __future__ import annotations

import json
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl

from .llm_client import LLMResponse, call_openrouter
from .prompt_templates import (
    PROMPT_STAGE2_ASSIGNMENT,
    optional_context_line,
    render_prompt,
)
from .stage1_discovery import QuestionContext, load_question_and_responses
from .themes import normalize_label


@dataclass
class ThemeInfo:
    theme_id: str
    theme_label: str
    theme_definition: str
    theme_group_label: str | None = None


@dataclass
class AssignmentResult:
    response_id: str
    themes: List[Tuple[str, float]]  # List of (theme_id, confidence) tuples
    overflow: bool = False
    no_match: bool = False
    input_tokens: int = 0
    output_tokens: int = 0
    processing_time: float = 0.0
    error: str | None = None


def load_theme_catalog(workbook_path: str, question_id: str) -> List[ThemeInfo]:
    """Load theme catalog filtered to active/candidate themes for the question."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    
    if "theme_catalog" not in wb.sheetnames:
        return []
    
    tsheet = wb["theme_catalog"]
    headers = [str(c).strip() for c in next(tsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
    
    # Also need theme_groups for group labels
    group_map: Dict[str, str] = {}
    if "theme_groups" in wb.sheetnames:
        gsheet = wb["theme_groups"]
        gheaders = [str(c).strip() for c in next(gsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        for row in gsheet.iter_rows(min_row=2, values_only=True):
            gdict = {gheaders[i]: row[i] for i in range(min(len(gheaders), len(row)))}
            gid = str(gdict.get("theme_group_id", "")).strip()
            glabel = str(gdict.get("theme_group_label", "")).strip()
            if gid:
                group_map[gid] = glabel
    
    themes: List[ThemeInfo] = []
    for row in tsheet.iter_rows(min_row=2, values_only=True):
        tdict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        
        # Filter by question_id and status
        t_question_id = str(tdict.get("question_id", "")).strip()
        status = str(tdict.get("status", "")).strip().lower()
        
        if t_question_id != question_id.strip():
            continue
        # Only include active and candidate (from discover) themes
        # Exclude: retired, candidate-add, candidate-retire (review suggestions)
        if status not in ("active", "candidate"):
            continue
        
        theme_id = str(tdict.get("theme_id", "")).strip()
        theme_label = str(tdict.get("theme_label", "")).strip()
        theme_definition = str(tdict.get("theme_definition", "")).strip()
        theme_group_id = str(tdict.get("theme_group_id", "")).strip()
        
        if theme_id and theme_label:
            themes.append(
                ThemeInfo(
                    theme_id=theme_id,
                    theme_label=theme_label,
                    theme_definition=theme_definition,
                    theme_group_label=group_map.get(theme_group_id),
                )
            )
    
    return themes


def _format_theme_catalog_block(theme_catalog: List[ThemeInfo]) -> str:
    lines: list[str] = []
    for theme in theme_catalog:
        group_part = f" (Group: {theme.theme_group_label})" if theme.theme_group_label else ""
        def_part = f" — {theme.theme_definition}" if theme.theme_definition else ""
        lines.append(f"- {theme.theme_id}: {theme.theme_label}{group_part}{def_part}")
    return "\n".join(lines)


def build_assignment_prompt(
    qctx: QuestionContext,
    theme_catalog: List[ThemeInfo],
    response_text: str,
    max_themes: int = 5,
) -> str:
    """Build prompt for theme assignment."""
    custom_context_section = optional_context_line("Custom context", qctx.custom_context)
    if custom_context_section:
        custom_context_section += "\n"

    return render_prompt(
        PROMPT_STAGE2_ASSIGNMENT,
        survey_name=qctx.survey_name,
        question_code=qctx.question_code,
        question_text=qctx.question_text,
        question_context_section=optional_context_line("Question context", qctx.question_context),
        custom_context_section=custom_context_section,
        theme_catalog_block=_format_theme_catalog_block(theme_catalog),
        response_text=response_text,
        max_themes=max_themes,
    )


def parse_assignment_json(text: str, valid_theme_ids: set[str]) -> Tuple[List[Tuple[str, float]], bool, bool]:
    """Parse assignment JSON response.
    
    Returns: (list of (theme_id, confidence) tuples, overflow flag, no_match flag)
    """
    # Try to extract JSON from markdown code blocks if present
    json_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
    if json_match:
        text = json_match.group(1)
    
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        # Try to find valid JSON object
        start_idx = text.find('{')
        if start_idx >= 0:
            for end_idx in range(len(text), start_idx, -1):
                try:
                    candidate = text[start_idx:end_idx] + '}'
                    data = json.loads(candidate)
                    break
                except json.JSONDecodeError:
                    continue
            else:
                raise ValueError("Could not parse LLM response as JSON")
        else:
            raise ValueError("Could not find JSON object in LLM response")
    
    themes_raw = data.get("themes") or []
    overflow = bool(data.get("overflow", False))
    
    themes: List[Tuple[str, float]] = []
    for item in themes_raw:
        if isinstance(item, dict):
            theme_id = str(item.get("theme_id", "")).strip()
            confidence = float(item.get("confidence", 1.0))
            if theme_id and theme_id in valid_theme_ids:
                themes.append((theme_id, confidence))
        elif isinstance(item, str):
            # Simple format: just theme_id
            theme_id = item.strip()
            if theme_id and theme_id in valid_theme_ids:
                themes.append((theme_id, 1.0))
    
    # Sort by confidence (highest first)
    themes.sort(key=lambda x: x[1], reverse=True)
    
    no_match = len(themes) == 0 and not overflow
    
    return themes, overflow, no_match


def assign_themes(
    workbook: str,
    *,
    question_id: str,
    model: str,
    max_themes: int = 5,
    sample_size: int | None = None,
    temperature: float = 0.2,
    max_tokens: int = 32000,
    response_ids: str | None = None,
    skip_existing: bool = False,
    verbose: int = 0,
    context_column: str | None = None,
) -> str:
    """Run Stage 2 theme assignment on a workbook.
    
    Args:
        max_themes: Maximum number of themes to assign per response (max: 50, default: 5).
    
    Returns the workbook path (same as input, modified in place).
    """
    # Validate max_themes (additional check - CLI also validates, but this protects direct function calls)
    if max_themes > 50:
        raise ValueError(f"max_themes cannot exceed 50 (got {max_themes})")
    
    start_time = time.time()
    
    # Load question and responses
    qctx, responses, total_responses = load_question_and_responses(
        workbook, question_id=question_id, sample_size=sample_size, context_column=context_column
    )
    
    # Apply response_ids filter if specified
    if response_ids:
        # Normalize target IDs to strings (handles both string and numeric input)
        target_ids = set()
        for rid in response_ids.split(','):
            rid = rid.strip()
            if not rid:
                continue
            # Normalize to string - if it's a number, convert to string representation
            try:
                # Try to normalize numeric IDs (e.g., "383" stays "383", but handles edge cases)
                normalized = str(int(float(rid)))
                target_ids.add(normalized)
            except ValueError:
                # Not a number, keep as string
                target_ids.add(rid)
        
        original_count = len(responses)
        # Get all available response IDs for diagnostic purposes (normalize to strings)
        available_ids = set()
        for r in responses:
            resp_id_raw = r.get("response_id") or r.get("id")
            if resp_id_raw is not None:
                # Normalize to string - handle both numeric and string IDs from Excel
                try:
                    # If it's a number in Excel, normalize it
                    normalized = str(int(float(resp_id_raw)))
                    available_ids.add(normalized)
                except (ValueError, TypeError):
                    # Not a number, convert to string
                    normalized = str(resp_id_raw).strip()
                    if normalized:
                        available_ids.add(normalized)
        
        # Filter responses - normalize IDs for comparison
        filtered_responses = []
        for r in responses:
            resp_id_raw = r.get("response_id") or r.get("id")
            if resp_id_raw is None:
                continue
            # Normalize response ID to string for comparison
            try:
                resp_id_normalized = str(int(float(resp_id_raw)))
            except (ValueError, TypeError):
                resp_id_normalized = str(resp_id_raw).strip()
            
            if resp_id_normalized and resp_id_normalized in target_ids:
                filtered_responses.append(r)
        
        responses = filtered_responses
        # Get found IDs for reporting (normalized)
        found_ids = set()
        for r in responses:
            resp_id_raw = r.get("response_id") or r.get("id")
            if resp_id_raw is not None:
                try:
                    found_ids.add(str(int(float(resp_id_raw))))
                except (ValueError, TypeError):
                    found_ids.add(str(resp_id_raw).strip())
        missing_ids = target_ids - found_ids
        if missing_ids and verbose >= 1:
            print(f"Warning: Response IDs not found: {', '.join(sorted(missing_ids))}", file=sys.stderr)
        if not responses:
            # Provide helpful diagnostic information
            error_msg = f"No matching responses found for specified response_ids. Requested: {', '.join(sorted(target_ids))}"
            if available_ids:
                # Show sample of available IDs to help user
                sample_ids = sorted(list(available_ids))[:10]
                error_msg += f"\nAvailable response IDs (sample of {min(10, len(available_ids))}): {', '.join(sample_ids)}"
                if len(available_ids) > 10:
                    error_msg += f" (and {len(available_ids) - 10} more...)"
            else:
                error_msg += f"\nNo response IDs found in workbook for question_id '{question_id}'"
            raise ValueError(error_msg)
        if verbose >= 1:
            print(f"Filtered to {len(responses)} responses (from {original_count}) using --response-ids", file=sys.stderr)
    
    # Load theme catalog
    theme_catalog = load_theme_catalog(workbook, question_id)
    if not theme_catalog:
        raise ValueError(f"No active/candidate themes found for question_id '{question_id}'")
    
    valid_theme_ids = {t.theme_id for t in theme_catalog}
    
    if verbose >= 1:
        print(f"Loaded {len(theme_catalog)} themes from catalog", file=sys.stderr)
        print(f"Processing {len(responses)} responses...", file=sys.stderr)
    
    # Load workbook for writing
    wb = openpyxl.load_workbook(workbook)
    rsheet = wb["responses_coded"]
    
    # Get headers
    headers = [str(c).strip() for c in next(rsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
    
    # Normalize model name for column names
    model_normalized = normalize_label(model)
    
    # Prepare column names
    theme_cols = [f"model_{model_normalized}_theme_{i}" for i in range(1, max_themes + 1)]
    confidence_cols = [f"model_{model_normalized}_confidence_{i}" for i in range(1, max_themes + 1)]
    truncated_col = f"model_{model_normalized}_truncated"
    
    # Add columns if needed (themes, then confidences, then truncated)
    for col in theme_cols + confidence_cols + [truncated_col]:
        if col not in headers:
            headers.append(col)
            rsheet.cell(row=1, column=len(headers), value=col)
    
    # Create header index map
    header_idx = {h: i + 1 for i, h in enumerate(headers)}
    
    # Build mapping from (response_id, response_text) to row_idx
    # This allows us to write assignments to the correct row when multiple rows share the same response_id
    response_to_row: Dict[Tuple[str, str], int] = {}
    response_id_to_rows: Dict[str, List[int]] = {}  # Keep for skip_existing logic (any row with that response_id)
    
    for row_idx, row in enumerate(rsheet.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) >= len(headers):
            resp_dict = {headers[i]: row[i] for i in range(len(headers))}
            resp_id = str(resp_dict.get("response_id", "")).strip()
            resp_qid = str(resp_dict.get("question_id", "")).strip()
            resp_text = str(resp_dict.get("response_text", "")).strip()
            
            if resp_id and resp_qid == question_id.strip():
                # Map (response_id, response_text) to row_idx for precise matching
                key = (resp_id, resp_text)
                response_to_row[key] = row_idx
                
                # Also track all rows for this response_id (for skip_existing and response-ids filtering)
                if resp_id not in response_id_to_rows:
                    response_id_to_rows[resp_id] = []
                response_id_to_rows[resp_id].append(row_idx)
    
    # Process each response
    results: List[AssignmentResult] = []
    total_input_tokens = 0
    total_output_tokens = 0
    skipped_count = 0
    
    for idx, response in enumerate(responses, 1):
        resp_id = str(response.get("response_id") or response.get("id") or "").strip()
        resp_text = str(response.get("response_text", "")).strip()
        
        if not resp_text or not resp_id:
            continue
        
        # Check skip_existing if enabled
        # Check if THIS specific row (matching both response_id and response_text) already has assignments
        if skip_existing:
            row_key = (resp_id, resp_text)
            if row_key in response_to_row:
                row_idx = response_to_row[row_key]
                theme_1_col_idx = header_idx.get(theme_cols[0])
                if theme_1_col_idx:
                    existing_value = rsheet.cell(row=row_idx, column=theme_1_col_idx).value
                    if existing_value and str(existing_value).strip():
                        skipped_count += 1
                        if verbose >= 1:
                            print(f"Skipping response {idx}/{len(responses)} (response_id: {resp_id}) - already assigned", file=sys.stderr)
                        continue
        
        if verbose >= 1:
            print(f"Processing response {idx}/{len(responses)} (response_id: {resp_id})...", file=sys.stderr, end="")
        
        prompt = build_assignment_prompt(qctx, theme_catalog, resp_text, max_themes)
        
        if verbose >= 2 and idx == 1:
            print("\n--- Assignment Prompt (first response) ---", file=sys.stderr)
            print(prompt, file=sys.stderr)
            print("--- End Prompt ---\n", file=sys.stderr)
        
        call_start = time.time()
        try:
            # Retry logic for empty responses (separate from network retries in call_openrouter)
            max_retries_empty = 2  # Retry up to 2 times for empty responses
            retry_delay_empty = 3.0  # Fixed delay in seconds (not exponential, since it's not a network error)
            llm_resp: LLMResponse | None = None
            
            for retry_attempt in range(max_retries_empty + 1):
                llm_resp = call_openrouter(
                    model=model,
                    prompt=prompt,
                    temperature=temperature,
                    max_tokens=max_tokens,
                )
                
                # Check if response has content
                if llm_resp.text:
                    # Success - break retry loop
                    break
                
                # Empty response - check finish_reason
                finish_reason = llm_resp.raw.get("choices", [{}])[0].get("finish_reason", "unknown")
                
                if retry_attempt < max_retries_empty:
                    # Retry empty response
                    if verbose >= 1:
                        print(f" Warning: Empty response (finish_reason: {finish_reason}), retrying ({retry_attempt + 1}/{max_retries_empty})...", file=sys.stderr, end="")
                    time.sleep(retry_delay_empty)
                    continue
                else:
                    # Final attempt failed
                    raise ValueError(
                        f"LLM returned empty response after {max_retries_empty + 1} attempts. "
                        f"Finish reason: {finish_reason}. "
                        f"This may indicate content filtering, model issues, or rate limiting."
                    )
            
            # Extract tokens (only count the final successful attempt)
            usage = llm_resp.raw.get("usage") or {}
            input_tokens = usage.get("prompt_tokens", 0)
            output_tokens = usage.get("completion_tokens", 0)
            total_input_tokens += input_tokens
            total_output_tokens += output_tokens
            
            # Final check (should not be needed if retry logic works, but safety check)
            if not llm_resp.text:
                raise ValueError("LLM returned empty response")
            
            # Parse response
            themes, overflow, no_match = parse_assignment_json(llm_resp.text, valid_theme_ids)
            
            # Check if overflow should be set (either flag set OR more than max_themes returned)
            original_count = len(themes)
            has_overflow = overflow or original_count > max_themes
            
            # Take top max_themes (sorted by confidence already)
            themes = themes[:max_themes]
            
            processing_time = time.time() - call_start
            
            result = AssignmentResult(
                response_id=resp_id,
                themes=themes,
                overflow=has_overflow,
                no_match=no_match,
                input_tokens=input_tokens,
                output_tokens=output_tokens,
                processing_time=processing_time,
            )
            
            if verbose >= 1:
                overflow_marker = " (overflow)" if has_overflow else ""
                print(f" {len(themes)} themes{overflow_marker}, {input_tokens}+{output_tokens} tokens, {processing_time:.2f}s", file=sys.stderr)
            
        except Exception as e:
            processing_time = time.time() - call_start
            result = AssignmentResult(
                response_id=resp_id,
                themes=[],
                error=str(e),
                processing_time=processing_time,
            )
            if verbose >= 1:
                print(f" ERROR: {e}", file=sys.stderr)
        
        results.append(result)
        
        # Write to workbook - find the correct row using both response_id and response_text
        row_key = (resp_id, resp_text)
        if row_key in response_to_row:
            row_idx = response_to_row[row_key]
            
            # First, clear all theme and confidence columns for this model to remove old data
            # This handles cases where max_themes changed or previous runs had different assignments
            for i in range(max_themes):
                if theme_cols[i] in header_idx:
                    theme_col_idx = header_idx[theme_cols[i]]
                    rsheet.cell(row=row_idx, column=theme_col_idx, value=None)
                if confidence_cols[i] in header_idx:
                    conf_col_idx = header_idx[confidence_cols[i]]
                    rsheet.cell(row=row_idx, column=conf_col_idx, value=None)
            
            # Write themes and confidence values
            for i, (theme_id, confidence) in enumerate(themes):
                if i < max_themes:
                    # Write theme_id
                    theme_col_idx = header_idx[theme_cols[i]]
                    rsheet.cell(row=row_idx, column=theme_col_idx, value=theme_id)
                    
                    # Write confidence (leave empty if missing/invalid)
                    conf_col_idx = header_idx[confidence_cols[i]]
                    conf_cell = rsheet.cell(row=row_idx, column=conf_col_idx)
                    try:
                        conf_value = float(confidence)
                        if 0.0 <= conf_value <= 1.0:
                            conf_cell.value = conf_value
                            # Set Excel number format (2 decimal places)
                            conf_cell.number_format = "0.00"
                        # If out of range, leave empty
                    except (ValueError, TypeError):
                        # If confidence is missing or invalid, leave empty
                        pass
            
            # Write truncated flag
            truncated_idx = header_idx[truncated_col]
            rsheet.cell(row=row_idx, column=truncated_idx, value="TRUE" if result.overflow else "FALSE")
    
    # Save workbook
    wb.save(workbook)
    
    # Log to model_runs
    created_ts = time.time()
    model_run_id = f"{normalize_label(qctx.survey_name)}_{normalize_label(qctx.question_code)}_{int(created_ts)}"
    
    if "model_runs" not in wb.sheetnames:
        mr_sheet = wb.create_sheet("model_runs")
        mr_sheet.append(["model_run_id", "created_ts", "stage", "model", "question_id", "temperature", "processed_count", "total_input_tokens", "total_output_tokens", "notes"])
    else:
        mr_sheet = wb["model_runs"]
    
    overflow_count = sum(1 for r in results if r.overflow)
    no_match_count = sum(1 for r in results if r.no_match)
    error_count = sum(1 for r in results if r.error is not None)
    
    mr_sheet.append([
        model_run_id,
        created_ts,
        "stage2_assignment",
        model,
        question_id,
        temperature,
        len(results),
        total_input_tokens,
        total_output_tokens,
        json.dumps({
            "overflow_count": overflow_count,
            "no_match_count": no_match_count,
            "error_count": error_count,
            "max_themes": max_themes,
        }),
    ])
    
    wb.save(workbook)
    
    total_duration = time.time() - start_time
    
    if verbose >= 1:
        invalid_count = sum(1 for r in results if any(tid not in valid_theme_ids for tid, _ in r.themes))
        print(f"\nAssignment complete:", file=sys.stderr)
        print(f"  Processed: {len(results)} responses", file=sys.stderr)
        if skip_existing and skipped_count > 0:
            print(f"  Skipped (already assigned): {skipped_count}", file=sys.stderr)
        print(f"  Overflow: {overflow_count}", file=sys.stderr)
        print(f"  No match: {no_match_count}", file=sys.stderr)
        print(f"  Invalid themes: {invalid_count}", file=sys.stderr)
        print(f"  Errors: {error_count}", file=sys.stderr)
        print(f"  Total time: {total_duration:.1f} seconds", file=sys.stderr)
        print(f"  Total tokens: {total_input_tokens} input + {total_output_tokens} output", file=sys.stderr)
    
    # Warn if skip_existing skipped all responses
    if skip_existing and skipped_count > 0 and len(results) == 0:
        print(f"Warning: All responses were skipped (already assigned). No new assignments were made.", file=sys.stderr)
    
    return workbook

