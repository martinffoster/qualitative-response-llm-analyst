from __future__ import annotations

import json
import random
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl

from .llm_client import LLMResponse, call_openrouter, sha256_text
from .prompt_templates import (
    PROMPT_STAGE1_DISCOVERY,
    format_responses_block,
    optional_context_line,
    render_prompt,
)
from .resources import bundled_template_path
from .themes import Theme, generate_theme_id, normalize_label
from .validation import validate_workbook


@dataclass
class QuestionContext:
    survey_name: str
    question_code: str
    question_text: str
    question_context: str | None = None
    research_context: str | None = None
    custom_context: str | None = None


def load_question_and_responses(
    workbook_path: str, question_id: str, sample_size: int | None = None, seed: int = 42, context_column: str | None = None
) -> Tuple[QuestionContext, List[Dict[str, Any]], int]:
    """Load question context and a sampled set of responses from workbook.

    Assumes template sheets: `question`, `responses_coded`.
    Finds the row in question sheet matching the given question_id.
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True)

    # Load question sheet and find matching question_id
    qsheet = wb["question"]
    headers_q = [str(c).strip() for c in next(qsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
    
    # Find the row with matching question_id
    qrec = None
    for row in qsheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers_q[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(min(len(headers_q), len(row)))}
        if row_dict.get("question_id", "").strip() == question_id.strip():
            qrec = row_dict
            break
    
    if qrec is None:
        raise ValueError(f"Question ID '{question_id}' not found in question sheet")

    survey_name = qrec.get("survey_name", "survey").strip() or "survey"
    question_code = qrec.get("question_code", question_id).strip() or question_id
    question_text = qrec.get("question_text", "").strip()
    if not question_text:
        raise ValueError(f"Question ID '{question_id}' has no question_text")
    question_context = qrec.get("question_context", "").strip() or None
    research_context = qrec.get("research_context", "").strip() or None
    custom_context = qrec.get(context_column, "").strip() or None if context_column else None

    qctx = QuestionContext(
        survey_name=survey_name,
        question_code=question_code,
        question_text=question_text,
        question_context=question_context,
        research_context=research_context,
        custom_context=custom_context,
    )

    # Load responses matching this question_id
    rsheet = wb["responses_coded"]
    headers_r = [str(c).strip() for c in next(rsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
    records: List[Dict[str, Any]] = []
    for row in rsheet.iter_rows(min_row=2, values_only=True):
        rec = {headers_r[i]: row[i] for i in range(min(len(headers_r), len(row)))}
        # Filter to responses matching this question_id
        rec_question_id = str(rec.get("question_id", "")).strip()
        if rec_question_id != question_id.strip():
            continue
        # Heuristic: require a response_text column
        txt = str(rec.get("response_text", "") or "").strip()
        if txt:
            records.append(rec)

    total_responses = len(records)
    random.Random(seed).shuffle(records)
    if sample_size is not None and sample_size > 0:
        records = records[:sample_size]
    return qctx, records, total_responses


def build_stage1_prompt(qctx: QuestionContext, responses: List[Dict[str, Any]], max_themes: int = 30) -> str:
    """Compose the discovery prompt that asks for themes and groups in JSON."""
    return render_prompt(
        PROMPT_STAGE1_DISCOVERY,
        survey_name=qctx.survey_name,
        question_code=qctx.question_code,
        question_text=qctx.question_text,
        question_context_section=optional_context_line("Question context", qctx.question_context),
        research_context_section=optional_context_line("Research context", qctx.research_context),
        custom_context_section=optional_context_line("Custom context", qctx.custom_context),
        responses_block=format_responses_block(responses),
        max_themes=max_themes,
    )


def parse_discovery_json(text: str) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    """Parse the LLM JSON for theme_groups and themes with basic validation.
    
    Attempts to extract valid JSON even if truncated. Looks for JSON block in markdown
    code fences if present.
    """
    # Try to extract JSON from markdown code blocks if present
    json_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
    if json_match:
        text = json_match.group(1)
    
    # Try to parse as-is first
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        # If parsing fails, try to find a valid JSON object starting from the first {
        start_idx = text.find('{')
        if start_idx >= 0:
            # Try to find a closing brace that creates valid JSON
            # This is a heuristic - we'll try progressively shorter substrings
            for end_idx in range(len(text), start_idx, -1):
                try:
                    candidate = text[start_idx:end_idx] + '}'
                    data = json.loads(candidate)
                    break
                except json.JSONDecodeError:
                    continue
            else:
                raise ValueError("Could not parse LLM response as JSON")
        else:
            raise ValueError("Could not find JSON object in LLM response")
    
    groups = data.get("theme_groups") or []
    themes = data.get("themes") or []

    norm_groups: List[Dict[str, str]] = []
    for g in groups:
        glabel = str(g.get("group_label", "")).strip()
        gdef = str(g.get("group_definition", "")).strip()
        if glabel:
            norm_groups.append({"group_label": glabel, "group_definition": gdef})

    norm_themes: List[Dict[str, str]] = []
    for t in themes:
        tlabel = str(t.get("theme_label", "")).strip()
        tdef = str(t.get("theme_definition", "")).strip()
        glabel = str(t.get("group_label", "")).strip()
        if tlabel:
            norm_themes.append(
                {
                    "theme_label": tlabel,
                    "theme_definition": tdef,
                    "group_label": glabel,
                }
            )
    return norm_groups, norm_themes


def write_discovery_results(
    *,
    input_workbook: str,
    output_workbook: str,
    qctx: QuestionContext,
    groups: List[Dict[str, str]],
    themes: List[Dict[str, str]],
    model_run_meta: Dict[str, Any],
) -> None:
    """Write theme_groups, theme_catalog, and an audit row in model_runs to a new workbook copy."""
    # Start from a copy of input workbook
    src_wb = openpyxl.load_workbook(input_workbook)

    # theme_groups sheet
    if "theme_groups" not in src_wb.sheetnames:
        ws_g = src_wb.create_sheet("theme_groups")
        headers_g = [
            "theme_group_id", "theme_group_label", "theme_group_definition",
            "question_id", "seen_years", "status", "last_edited_by", "last_edited_at", "notes"
        ]
        ws_g.append(headers_g)
        next_row = 2
    else:
        ws_g = src_wb["theme_groups"]
        # Get headers to map column positions
        headers_g = [str(c).strip() for c in next(ws_g.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        # Find the last row with actual data (skip empty rows)
        next_row = 2  # Start at row 2 (after header)
        for row_idx, row in enumerate(ws_g.iter_rows(min_row=2, values_only=True), start=2):
            # Check if row has any non-empty data in first 3 columns (id, label, definition)
            if row and any(cell is not None and str(cell).strip() for cell in (row[0:3] if len(row) >= 3 else row)):
                next_row = row_idx + 1  # Next row after this non-empty row
            else:
                break

    # Build group_id mapping
    group_label_to_id: Dict[str, str] = {}
    next_group_seq = 1
    existing_labels = set()
    
    for row in ws_g.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < len(headers_g):
            continue
        # Skip empty rows
        if not any(cell is not None and str(cell).strip() for cell in (row[0:3] if len(row) >= 3 else row)):
            continue
        # Map row to dict for easier access
        row_dict = {headers_g[i]: row[i] if i < len(row) else None for i in range(len(headers_g))}
        label = (row_dict.get("theme_group_label") or "").strip()
        if label:
            existing_labels.add(label)

    # Prepare metadata for new rows
    current_year = datetime.now().year
    model_name = model_run_meta.get("model", "unknown")
    current_timestamp = datetime.now()

    for g in groups:
        glabel = g["group_label"].strip()
        gdef = g.get("group_definition", "").strip()
        if not glabel:
            continue
        if glabel in existing_labels:
            # find existing id by rescanning; cheap for small tables
            for row in ws_g.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= len(headers_g):
                    row_dict = {headers_g[i]: row[i] if i < len(row) else None for i in range(len(headers_g))}
                    if any(cell is not None and str(cell).strip() for cell in (row[0:3] if len(row) >= 3 else row)):
                        if (row_dict.get("theme_group_label") or "").strip() == glabel:
                            group_label_to_id[glabel] = (row_dict.get("theme_group_id") or "").strip()
                            break
            continue
        gid = generate_theme_id(qctx.survey_name, qctx.question_code, glabel, next_group_seq)
        next_group_seq += 1
        group_label_to_id[glabel] = gid
        
        # Write all columns according to schema
        col_map = {h: i + 1 for i, h in enumerate(headers_g)}
        ws_g.cell(row=next_row, column=col_map.get("theme_group_id", 1), value=gid)
        ws_g.cell(row=next_row, column=col_map.get("theme_group_label", 2), value=glabel)
        ws_g.cell(row=next_row, column=col_map.get("theme_group_definition", 3), value=gdef)
        ws_g.cell(row=next_row, column=col_map.get("question_id", 4), value=qctx.question_code)
        ws_g.cell(row=next_row, column=col_map.get("seen_years", 5), value=current_year)
        ws_g.cell(row=next_row, column=col_map.get("status", 6), value="candidate")
        ws_g.cell(row=next_row, column=col_map.get("last_edited_by", 7), value=model_name)
        ws_g.cell(row=next_row, column=col_map.get("last_edited_at", 8), value=current_timestamp)
        # notes column (9) - leave empty
        next_row += 1

    # theme_catalog sheet
    if "theme_catalog" not in src_wb.sheetnames:
        ws_t = src_wb.create_sheet("theme_catalog")
        headers_t = [
            "theme_id", "theme_group_id", "theme_label", "theme_definition",
            "question_id", "seen_years", "status", "last_edited_by", "last_edited_at", "replaced_prior_theme_ids"
        ]
        ws_t.append(headers_t)
        next_theme_row = 2
    else:
        ws_t = src_wb["theme_catalog"]
        # Get headers to map column positions
        headers_t = [str(c).strip() for c in next(ws_t.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        # Find the last row with actual data (skip empty rows)
        next_theme_row = 2  # Start at row 2 (after header)
        for row_idx, row in enumerate(ws_t.iter_rows(min_row=2, values_only=True), start=2):
            # Check if row has any non-empty data in first 3 columns (id, group_id, label)
            if row and any(cell is not None and str(cell).strip() for cell in (row[0:3] if len(row) >= 3 else row)):
                next_theme_row = row_idx + 1  # Next row after this non-empty row
            else:
                break

    existing_theme_labels = set()
    for row in ws_t.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < len(headers_t):
            continue
        # Skip empty rows
        if not any(cell is not None and str(cell).strip() for cell in (row[0:3] if len(row) >= 3 else row)):
            continue
        # Map row to dict for easier access
        row_dict = {headers_t[i]: row[i] if i < len(row) else None for i in range(len(headers_t))}
        label = (row_dict.get("theme_label") or "").strip()
        if label:
            existing_theme_labels.add(label)

    next_theme_seq = 1
    discovered_theme_ids: List[str] = []
    for t in themes:
        tlabel = t["theme_label"].strip()
        tdef = t.get("theme_definition", "").strip()
        glabel = t.get("group_label", "").strip()
        if not tlabel:
            continue
        if tlabel in existing_theme_labels:
            # skip duplicates by label
            continue
        gid = group_label_to_id.get(glabel) if glabel else None
        tid = generate_theme_id(qctx.survey_name, qctx.question_code, tlabel, next_theme_seq)
        next_theme_seq += 1
        
        # Write all columns according to schema
        col_map = {h: i + 1 for i, h in enumerate(headers_t)}
        ws_t.cell(row=next_theme_row, column=col_map.get("theme_id", 1), value=tid)
        ws_t.cell(row=next_theme_row, column=col_map.get("theme_group_id", 2), value=gid)
        ws_t.cell(row=next_theme_row, column=col_map.get("theme_label", 3), value=tlabel)
        ws_t.cell(row=next_theme_row, column=col_map.get("theme_definition", 4), value=tdef)
        ws_t.cell(row=next_theme_row, column=col_map.get("question_id", 5), value=qctx.question_code)
        ws_t.cell(row=next_theme_row, column=col_map.get("seen_years", 6), value=current_year)
        ws_t.cell(row=next_theme_row, column=col_map.get("status", 7), value="candidate")
        ws_t.cell(row=next_theme_row, column=col_map.get("last_edited_by", 8), value=model_name)
        ws_t.cell(row=next_theme_row, column=col_map.get("last_edited_at", 9), value=current_timestamp)
        # replaced_prior_theme_ids column (10) - leave empty
        next_theme_row += 1
        discovered_theme_ids.append(tid)

    # model_runs sheet
    if "model_runs" not in src_wb.sheetnames:
        ws_m = src_wb.create_sheet("model_runs")
        ws_m.append(["model_run_id", "created_ts", "stage", "model", "temperature", "sample_size", "prompt_hash", "notes"])  # conservative schema
    else:
        ws_m = src_wb["model_runs"]

    ws_m.append(
        [
            model_run_meta.get("model_run_id"),
            model_run_meta.get("created_ts"),
            "stage1_discovery",
            model_run_meta.get("model"),
            model_run_meta.get("temperature"),
            model_run_meta.get("sample_size"),
            model_run_meta.get("prompt_hash"),
            json.dumps({"discovered_theme_ids": discovered_theme_ids}),
        ]
    )

    # Save to output path
    out_path = Path(output_workbook)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    src_wb.save(str(out_path))


def discover(
    workbook: str,
    *,
    question_id: str,
    model: str,
    sample_size: int | None = None,
    temperature: float = 0.2,
    max_tokens: int = 32000,
    max_themes: int = 30,
    output: str | None = None,
    verbose: int = 0,
    context_column: str | None = None,
) -> str:
    """Run Stage 1 discovery on a workbook and write results to a new file.

    Returns the output workbook path.
    """
    start_time = time.time()
    
    # Validate against template if available
    tpl_default = bundled_template_path()
    try:
        _valid, _issues = validate_workbook(workbook, tpl_default)
        # We don't exit on failure; discovery can still proceed on close-enough shapes
    except Exception:
        pass

    qctx, responses, total_responses = load_question_and_responses(workbook, question_id=question_id, sample_size=sample_size, context_column=context_column)
    
    if verbose >= 1:
        print(f"Sampling {len(responses)} responses from {total_responses} total...", file=sys.stderr)
        print(f"Calling {model} with discovery prompt...", file=sys.stderr)
    
    prompt = build_stage1_prompt(qctx, responses, max_themes=max_themes)
    
    if verbose >= 2:
        print("\n--- Discovery Prompt ---", file=sys.stderr)
        print(prompt, file=sys.stderr)
        print("--- End Prompt ---\n", file=sys.stderr)
    
    llm_call_start = time.time()
    llm_resp: LLMResponse = call_openrouter(
        model=model, prompt=prompt, temperature=temperature, max_tokens=max_tokens
    )
    llm_call_duration = time.time() - llm_call_start

    # Check if response was truncated
    finish_reason = None
    try:
        choices = llm_resp.raw.get("choices") or []
        if choices:
            finish_reason = choices[0].get("finish_reason")
    except Exception:
        pass
    
    if not llm_resp.text:
        # Check for reasoning-based models that might have content elsewhere
        raise ValueError(
            f"LLM returned empty response. Finish reason: {finish_reason}. "
            f"This may indicate the response was truncated or the model format isn't supported."
        )
    
    if finish_reason == "length":
        print(
            f"Warning: LLM response was truncated (hit max_tokens limit). Results may be incomplete.",
            file=sys.stderr,
        )

    if verbose >= 2:
        print("\n--- LLM Response ---", file=sys.stderr)
        print(llm_resp.text, file=sys.stderr)
        print("--- End Response ---\n", file=sys.stderr)

    # Parse LLM JSON; if parsing fails, raise for visibility
    groups, themes = parse_discovery_json(llm_resp.text)

    created_ts = time.time()
    model_run_id = f"{normalize_label(qctx.survey_name)}_{normalize_label(qctx.question_code)}_{int(created_ts)}"
    
    # Extract token usage from response
    input_tokens = 0
    output_tokens = 0
    usage = llm_resp.raw.get("usage") or {}
    if usage:
        input_tokens = usage.get("prompt_tokens", 0)
        output_tokens = usage.get("completion_tokens", 0)
    
    model_run_meta = {
        "model_run_id": model_run_id,
        "created_ts": created_ts,
        "model": llm_resp.model,
        "temperature": temperature,
        "sample_size": sample_size if sample_size is not None else 0,
        "prompt_hash": llm_resp.prompt_hash,
    }

    in_path = Path(workbook)
    if output is None:
        output = str(in_path.with_suffix("").as_posix() + f".discovered.{normalize_label(llm_resp.model)}.xlsx")

    write_discovery_results(
        input_workbook=workbook,
        output_workbook=output,
        qctx=qctx,
        groups=groups,
        themes=themes,
        model_run_meta=model_run_meta,
    )

    total_duration = time.time() - start_time
    
    if verbose >= 1:
        print(
            f"Discovery complete: {len(groups)} groups, {len(themes)} themes found in {total_duration:.1f} seconds "
            f"({input_tokens} input + {output_tokens} output tokens)",
            file=sys.stderr,
        )
    
    return output




