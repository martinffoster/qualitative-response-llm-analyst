from __future__ import annotations

import re
import sys
from collections import Counter
from dataclasses import dataclass
from typing import Any, Dict, List, Set, Tuple

import openpyxl
from openpyxl.chart import BarChart, Reference
import pandas as pd

from .stage1_discovery import QuestionContext, load_question_and_responses
from .stage2_assignment import ThemeInfo, load_theme_catalog


@dataclass
class ThemeCount:
    theme_id: str
    theme_label: str
    theme_group_id: str | None
    theme_group_label: str | None
    count: int


def detect_models_in_workbook(workbook_path: str, question_id: str) -> List[str]:
    """Detect all models that have assignment columns in responses_coded sheet.
    
    Returns list of normalized model names (e.g., ['openai_gpt_5_mini', 'google_gemini_2_5_flash']).
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    
    if "responses_coded" not in wb.sheetnames:
        return []
    
    rsheet = wb["responses_coded"]
    headers = [str(c).strip() for c in next(rsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
    
    # Pattern: model_<normalized_name>_theme_1
    model_pattern = re.compile(r'^model_(.+)_theme_1$')
    models: Set[str] = set()
    
    for header in headers:
        match = model_pattern.match(header)
        if match:
            models.add(match.group(1))
    
    wb.close()
    return sorted(list(models))


def count_themes_for_model(
    workbook_path: str, question_id: str, model_normalized: str
) -> Counter[str]:
    """Count occurrences of each theme_id in a model's assignment columns.
    
    Dynamically detects all theme columns for the model by scanning headers.
    Works with any number of theme columns (1, 5, 10, 50, etc.).
    
    Returns Counter mapping theme_id -> count.
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    rsheet = wb["responses_coded"]
    
    headers = [str(c).strip() for c in next(rsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
    
    # Dynamically find all theme column indices for this model
    # Pattern: model_{model_normalized}_theme_{N} where N is any number
    theme_pattern = re.compile(r'^model_' + re.escape(model_normalized) + r'_theme_(\d+)$')
    theme_col_indices: List[int] = []
    
    for i, header in enumerate(headers):
        match = theme_pattern.match(header)
        if match:
            theme_col_indices.append(i)
    
    # Sort by theme number (extract number from header for proper ordering)
    def get_theme_num(idx: int) -> int:
        header = headers[idx]
        match = theme_pattern.match(header)
        return int(match.group(1)) if match else 0
    
    theme_col_indices.sort(key=get_theme_num)
    
    if not theme_col_indices:
        wb.close()
        return Counter()
    
    question_id_idx = headers.index("question_id") if "question_id" in headers else None
    
    counts: Counter[str] = Counter()
    for row in rsheet.iter_rows(min_row=2, values_only=True):
        # Skip rows that are too short (don't have enough columns)
        if len(row) < len(headers):
            continue
        
        # Filter by question_id if column exists
        if question_id_idx is not None:
            row_question_id = str(row[question_id_idx] or "").strip()
            if row_question_id != question_id.strip():
                continue
        
        # Count themes from this model's columns
        for col_idx in theme_col_indices:
            theme_id = str(row[col_idx] or "").strip()
            if theme_id:
                counts[theme_id] += 1
    
    wb.close()
    return counts


def load_theme_catalog_with_groups(
    workbook_path: str, question_id: str
) -> Tuple[List[ThemeInfo], Dict[str, Dict[str, str]]]:
    """Load theme catalog and create a mapping of theme_id -> {theme_group_id, theme_group_label}.
    
    Returns: (theme_list, theme_to_group_map)
    """
    themes = load_theme_catalog(workbook_path, question_id)
    
    # Load theme_group_id from theme_catalog sheet directly
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    theme_to_group: Dict[str, Dict[str, str]] = {}
    
    if "theme_catalog" in wb.sheetnames:
        tsheet = wb["theme_catalog"]
        headers = [str(c).strip() for c in next(tsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        
        for row in tsheet.iter_rows(min_row=2, values_only=True):
            tdict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            
            t_question_id = str(tdict.get("question_id", "")).strip()
            status = str(tdict.get("status", "")).strip().lower()
            theme_id = str(tdict.get("theme_id", "")).strip()
            theme_group_id = str(tdict.get("theme_group_id", "")).strip() or None
            
            if t_question_id != question_id.strip():
                continue
            if status not in ("active", "candidate"):
                continue
            if theme_id:
                theme_to_group[theme_id] = {
                    "theme_group_id": theme_group_id,
                    "theme_group_label": None,  # Will be filled from group lookup
                }
    
    wb.close()
    return themes, theme_to_group


def create_summary_sheet(
    workbook_path: str,
    question_id: str,
    model_normalized: str,
    theme_catalog: List[ThemeInfo],
    theme_counts: Counter[str],
    theme_to_group: Dict[str, Dict[str, str]],
    group_lookup: Dict[str, str],
    verbose: int = 0,
) -> Tuple[int, int, int]:
    """Create summary sheet for a model.
    
    Returns: (total_themes_in_catalog, consistency_issues_count, orphaned_theme_count)
    """
    wb = openpyxl.load_workbook(workbook_path)
    sheet_name = f"summary_{model_normalized}"
    
    # Delete existing sheet if present
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    sheet = wb.create_sheet(sheet_name)
    
    # Create theme_id -> ThemeInfo lookup
    theme_lookup: Dict[str, ThemeInfo] = {t.theme_id: t for t in theme_catalog}
    
    # Check for orphaned theme_ids (themes assigned in responses but not in catalog)
    orphaned_themes = [tid for tid in theme_counts.keys() if tid not in theme_lookup]
    orphaned_count = len(orphaned_themes)
    
    if orphaned_count > 0 and verbose >= 1:
        print(
            f"  Warning: Found {orphaned_count} theme_id(s) in assignments not in catalog: {', '.join(orphaned_themes[:5])}"
            + (f" (and {orphaned_count - 5} more)" if orphaned_count > 5 else ""),
            file=sys.stderr,
        )
    
    # Build summary data
    summary_rows: List[ThemeCount] = []
    consistency_issues = 0
    
    for theme in theme_catalog:
        count = theme_counts.get(theme.theme_id, 0)
        
        # Get theme_group_id from theme_to_group map
        group_info = theme_to_group.get(theme.theme_id, {})
        group_id = group_info.get("theme_group_id")
        group_label = None
        
        # Check for consistency: if theme has a group_id, verify it exists in group_lookup
        if group_id:
            group_label = group_lookup.get(group_id)
            if group_label is None and verbose >= 1:
                print(
                    f"  Warning: theme_id '{theme.theme_id}' references theme_group_id '{group_id}' "
                    f"not found in theme_groups sheet",
                    file=sys.stderr,
                )
                consistency_issues += 1
        else:
            # No group_id - this is OK, themes don't need to be grouped
            pass
        
        summary_rows.append(
            ThemeCount(
                theme_id=theme.theme_id,
                theme_label=theme.theme_label,
                theme_group_id=group_id,
                theme_group_label=group_label,
                count=count,
            )
        )
    
    # Sort by count descending
    summary_rows.sort(key=lambda x: x.count, reverse=True)
    
    # Write header
    sheet.append(["theme_group_id", "theme_group_label", "theme_id", "theme_label", "theme_id_count"])
    
    # Write data
    for row in summary_rows:
        sheet.append([
            row.theme_group_id or "",
            row.theme_group_label or "",
            row.theme_id,
            row.theme_label,
            row.count,
        ])
    
    wb.save(workbook_path)
    
    return len(theme_catalog), consistency_issues, orphaned_count


def create_pivot_and_chart(
    workbook_path: str,
    model_normalized: str,
    summary_sheet_name: str,
    chart_type: str,
    theme_to_group: Dict[str, Dict[str, str]],
    group_lookup: Dict[str, str],
    verbose: int = 0,
) -> None:
    """Create pivot table structure and chart for a model.
    
    chart_type: 'stackedbar' (requires groups), 'bar' (ignores groups), or 'auto'
    """
    wb = openpyxl.load_workbook(workbook_path)
    
    if summary_sheet_name not in wb.sheetnames:
        raise ValueError(f"Summary sheet '{summary_sheet_name}' not found")
    
    summary_sheet = wb[summary_sheet_name]
    
    # Read summary data into pandas
    data = []
    headers = [str(c).strip() for c in next(summary_sheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
    for row in summary_sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        data.append(row_dict)
    
    df = pd.DataFrame(data)
    
    # Determine chart type
    has_groups = bool(group_lookup) and df["theme_group_label"].notna().any()
    
    if chart_type == "auto":
        chart_type = "stackedbar" if has_groups else "bar"
    
    if chart_type == "stackedbar":
        if not has_groups:
            if verbose >= 1:
                print(f"  Chart type 'stackedbar' requires theme groups, but none found. Skipping chart creation.", file=sys.stderr)
            wb.close()
            return
        # Create pivot: groups as rows (categories), themes as columns (stacked series)
        pivot_df = df.pivot_table(
            values="theme_id_count",
            index="theme_group_label",
            columns="theme_label",
            aggfunc="sum",
            fill_value=0
        )
        # Add sum column
        pivot_df["Total"] = pivot_df.sum(axis=1)
        # Sort by total ascending (smallest to largest) so chart shows largest at top (Excel reverse)
        pivot_df = pivot_df.sort_values("Total", ascending=True)
    else:  # chart_type == "bar"
        # Simple bar chart: theme_label vs count
        # Filter out zero counts for cleaner chart
        pivot_df = df[["theme_label", "theme_id_count"]].copy()
        pivot_df = pivot_df[pivot_df["theme_id_count"] > 0]
        # Sort descending by count, but write ascending (for Excel quirk)
        pivot_df = pivot_df.sort_values("theme_id_count", ascending=False)
        # Reset index for writing
        pivot_df = pivot_df.reset_index(drop=True)
        # Re-sort ascending for Excel quirk (chart will show highest at top)
        pivot_df = pivot_df.sort_values("theme_id_count", ascending=True).reset_index(drop=True)
    
    # Write pivot table to sheet
    pivot_sheet_name = f"pivot_{model_normalized}"
    if pivot_sheet_name in wb.sheetnames:
        wb.remove(wb[pivot_sheet_name])
    pivot_sheet = wb.create_sheet(pivot_sheet_name)
    
    # Write header row: first column is empty for row labels, then data columns
    if chart_type == "stackedbar":
        # Header: first col with label for categories, then theme labels (series names), then Total
        pivot_sheet.cell(row=1, column=1, value="Theme Group")  # Header for category column
        # Write theme labels (all columns except Total)
        theme_cols = [col for col in pivot_df.columns if col != "Total"]
        for col_idx, theme_label in enumerate(theme_cols, start=2):
            pivot_sheet.cell(row=1, column=col_idx, value=str(theme_label))
        # Write Total column header
        total_col = len(theme_cols) + 2
        pivot_sheet.cell(row=1, column=total_col, value="Total")
        
        # Write data: groups as rows, theme counts as columns, then Total
        for row_idx, (group_label, row_data) in enumerate(pivot_df.iterrows(), start=2):
            pivot_sheet.cell(row=row_idx, column=1, value=str(group_label))
            # Write theme values (excluding Total)
            for col_idx, theme_label in enumerate(theme_cols, start=2):
                pivot_sheet.cell(row=row_idx, column=col_idx, value=float(row_data[theme_label]))
            # Write Total value
            pivot_sheet.cell(row=row_idx, column=total_col, value=float(row_data["Total"]))
    else:  # bar
        # Header: theme_label, theme_id_count
        pivot_sheet.cell(row=1, column=1, value="Theme")
        pivot_sheet.cell(row=1, column=2, value="Count")
        
        # Write data: theme labels and counts
        for row_idx, (_, row_data) in enumerate(pivot_df.iterrows(), start=2):
            pivot_sheet.cell(row=row_idx, column=1, value=str(row_data["theme_label"]))
            pivot_sheet.cell(row=row_idx, column=2, value=float(row_data["theme_id_count"]))
    
    # Create chart sheet
    chart_sheet_name = f"chart_{model_normalized}"
    if chart_sheet_name in wb.sheetnames:
        wb.remove(wb[chart_sheet_name])
    chart_sheet = wb.create_sheet(chart_sheet_name)
    
    # Create chart
    chart = BarChart()
    chart.type = "bar"  # horizontal bars (categories on Y-axis, values on X-axis)
    chart.title = f"Theme Assignments - {model_normalized}"
    chart.x_axis.title = "Theme" if chart_type == "bar" else "Theme Group"
    chart.y_axis.title = "Count"
    
    if chart_type == "stackedbar":
        chart.grouping = "stacked"
        chart.overlap = 100
        
        # Categories: theme groups (rows, column 1) - these appear on Y-axis
        # Data series: one per theme (columns, starting at column 2, excluding Total) - these stack within each group bar
        num_rows = len(pivot_df)
        # Count theme columns (exclude Total which is at the end)
        theme_cols_count = len([col for col in pivot_df.columns if col != "Total"])

        # 1) ADD SERIES FIRST (include header row so titles_from_data picks up theme names)
        #   - why? In openpyxl, categories are attached per-series. Calling chart.set_categories(categories) 
        #          before you add any series doesn’t bind the labels to the series, so Excel falls back to 1,2,3…
        #
        # Data series: one per theme (columns 2..theme_cols_count+1, excluding Total)
        # These become the stacked segments and appear in the legend
        # Add all data series at once - each theme column becomes a series in the legend
        # Reference includes header row (min_row=1) so titles_from_data=True picks up theme names
        for col_idx in range(2, theme_cols_count + 2):              # columns B..(B+theme_cols_count-1)
            data_ref = Reference(pivot_sheet, min_col=col_idx, min_row=1, max_row=num_rows + 1)
            chart.add_data(data_ref, titles_from_data=True)

        # 2) THEN SET CATEGORIES (EXCLUDE HEADER ROW)
        # Categories: group labels (column 1, rows 2 to num_rows+1) - these become Y-axis labels
        # For horizontal bar charts (type="bar"), categories go on Y-axis
        # Include row 1 in reference range so Excel recognizes column A as categories
        categories = Reference(pivot_sheet, min_col=1, min_row=1, max_row=num_rows + 1)
        chart.set_categories(categories)
        
        # 3) AXIS BEHAVIOUR
        # Ensure Y-axis displays category labels properly
        # For horizontal bar charts, Y-axis is the category axis
        chart.y_axis.delete = False  # Prevent axis deletion (openpyxl 3.1.4+ issue)
        chart.y_axis.majorTickMark = "out"
        chart.y_axis.tickLblPos = "low"  # Position labels at low position (left side for horizontal bars)
        chart.y_axis.reverseOrder = False  # Uncheck "Values in reverse order" in Excel
        

        
        # Force chart to recognize column A as categories by ensuring it's part of the structure
        # The categories reference already points to column A, but we need to make sure
        # Excel can see it when auto-detecting the data range
        
        # Configure axes
        # X-axis: values (count) - ensure it starts at 0 and goes left to right
        chart.x_axis.scaling.min = 0
        chart.x_axis.delete = False  # Prevent axis deletion (openpyxl 3.1.4+ issue)
        # Y-axis: categories (theme groups)
        # Don't use maxMin orientation - it causes "Values in reverse order" to be checked
        # Instead, we'll sort the data in descending order if we want highest at top
        # chart.y_axis.scaling.orientation = "maxMin"  # Removed - causes reverse order checkbox
        
        # Position legend to the right without overlapping chart area
        chart.legend.position = "r"  # right side
        chart.legend.overlay = False  # don't overlay on chart (positioned separately)
    else:  # bar
        # Simple bar chart: theme_label vs count
        num_rows = len(pivot_df)
        
        # Categories: theme labels (column 1, rows 2 to num_rows+1)
        categories = Reference(pivot_sheet, min_col=1, min_row=2, max_row=num_rows + 1)
        chart.set_categories(categories)
        
        # Data: count column (column 2, rows 1 to num_rows+1)
        data_ref = Reference(pivot_sheet, min_col=2, min_row=1, max_row=num_rows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        
        # Reverse Y axis to show highest count at top
        chart.y_axis.scaling.orientation = "maxMin"
    
    # Set chart size and add chart to sheet
    chart.width = 30     # width in inches (~30 = full sheet width)
    chart.height = 20    # height in inches (~20 = full page height)
    chart_sheet.add_chart(chart, "B2")
    
    wb.save(workbook_path)
    wb.close()


def summarize(
    workbook: str,
    *,
    question_id: str,
    chart_type: str | None = None,
    verbose: int = 0,
) -> str:
    """Create summary sheets for each model found in responses_coded.
    
    Args:
        workbook: Path to workbook
        question_id: Question ID to process
        chart_type: Optional chart type: 'stackedbar', 'bar', or 'auto'
        verbose: Verbosity level
    
    Returns the workbook path (modified in place).
    """
    # Load theme catalog and theme-to-group mapping
    theme_catalog, theme_to_group = load_theme_catalog_with_groups(workbook, question_id)
    if not theme_catalog:
        raise ValueError(f"No active/candidate themes found for question_id '{question_id}'")
    
    # Load theme groups for lookup
    wb = openpyxl.load_workbook(workbook, read_only=True)
    group_lookup: Dict[str, str] = {}
    if "theme_groups" in wb.sheetnames:
        gsheet = wb["theme_groups"]
        gheaders = [str(c).strip() for c in next(gsheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c is not None]
        for row in gsheet.iter_rows(min_row=2, values_only=True):
            gdict = {gheaders[i]: row[i] for i in range(min(len(gheaders), len(row)))}
            gid = str(gdict.get("theme_group_id", "")).strip()
            glabel = str(gdict.get("theme_group_label", "")).strip()
            gqid = str(gdict.get("question_id", "")).strip()
            if gid and gqid == question_id.strip():
                group_lookup[gid] = glabel
    
    wb.close()
    
    if verbose >= 1:
        print(f"Loaded {len(theme_catalog)} themes from catalog", file=sys.stderr)
        if group_lookup:
            print(f"Loaded {len(group_lookup)} theme groups", file=sys.stderr)
        else:
            print(f"No theme groups found (themes may not be grouped)", file=sys.stderr)
    
    # Detect models
    models = detect_models_in_workbook(workbook, question_id)
    
    if not models:
        if verbose >= 1:
            print("No models detected in responses_coded (no model_*_theme_1 columns found)", file=sys.stderr)
        return workbook
    
    if verbose >= 1:
        print(f"Found {len(models)} model(s): {', '.join(models)}", file=sys.stderr)
    
    # Create summary sheet for each model
    total_consistency_issues = 0
    total_orphaned = 0
    
    for model in models:
        if verbose >= 1:
            print(f"\nCreating summary for model: {model}", file=sys.stderr)
        
        # Count themes
        theme_counts = count_themes_for_model(workbook, question_id, model)
        
        # Create summary sheet
        total_themes, consistency_issues, orphaned_count = create_summary_sheet(
            workbook, question_id, model, theme_catalog, theme_counts, theme_to_group, group_lookup, verbose
        )
        
        total_consistency_issues += consistency_issues
        total_orphaned += orphaned_count
        
        if verbose >= 1:
            assigned_count = sum(1 for count in theme_counts.values() if count > 0)
            total_assignments = sum(theme_counts.values())
            print(
                f"  Themes: {assigned_count}/{total_themes} assigned, "
                f"total assignments: {total_assignments}",
                file=sys.stderr,
            )
        
        # Create pivot table and chart if requested
        if chart_type:
            summary_sheet_name = f"summary_{model}"
            try:
                create_pivot_and_chart(
                    workbook,
                    model,
                    summary_sheet_name,
                    chart_type,
                    theme_to_group,
                    group_lookup,
                    verbose,
                )
                if verbose >= 1:
                    print(f"  Created pivot table and chart", file=sys.stderr)
            except Exception as e:
                if verbose >= 1:
                    print(f"  Warning: Could not create chart: {e}", file=sys.stderr)
    
    if verbose >= 1:
        print("\n" + "=" * 60, file=sys.stderr)
        if total_consistency_issues > 0:
            print(f"Total consistency warnings: {total_consistency_issues}", file=sys.stderr)
        if total_orphaned > 0:
            print(f"Total orphaned theme_ids: {total_orphaned}", file=sys.stderr)
        if total_consistency_issues == 0 and total_orphaned == 0:
            print("No data consistency issues found", file=sys.stderr)
        print("Summary sheets created successfully", file=sys.stderr)
    
    return workbook
