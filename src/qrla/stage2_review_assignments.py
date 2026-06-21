from __future__ import annotations

import re
import sys
from collections import Counter
from typing import Any, Dict, List, Tuple

import openpyxl

from .stage2_assignment import load_theme_catalog


def collect_theme_assignments_per_response(
    workbook_path: str, question_id: str
) -> List[Tuple[Dict[str, Any], List[Tuple[str, int]]]]:
    """Collect theme assignments per response, preserving original order.
    
    Returns: List of tuples (response_data, theme_assignments) where:
    - response_data = Dict with question_id, response_id, response_text
    - theme_assignments = List of tuples (theme_id, model_count) sorted by count descending
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    
    if "responses_coded" not in wb.sheetnames:
        wb.close()
        return []
    
    rsheet = wb["responses_coded"]
    headers = [str(c).strip() if c is not None else "" for c in next(rsheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    
    # Find base column indices
    question_id_idx = headers.index("question_id") if "question_id" in headers else None
    response_id_idx = headers.index("response_id") if "response_id" in headers else None
    response_text_idx = headers.index("response_text") if "response_text" in headers else None
    
    if question_id_idx is None or response_id_idx is None or response_text_idx is None:
        wb.close()
        raise ValueError("responses_coded sheet must have question_id, response_id, and response_text columns")
    
    # Find all theme column indices (pattern: model_*_theme_*)
    theme_pattern = re.compile(r'^model_.+_theme_\d+$')
    theme_col_indices: List[int] = []
    
    for i, header in enumerate(headers):
        if theme_pattern.match(header):
            theme_col_indices.append(i)
    
    if not theme_col_indices:
        wb.close()
        return []
    
    # Process each response row in original order
    results: List[Tuple[Dict[str, Any], List[Tuple[str, int]]]] = []
    
    for row in rsheet.iter_rows(min_row=2, values_only=True):
        # Skip rows that are too short
        if len(row) < len(headers):
            continue
        
        # Extract base response data
        row_question_id = str(row[question_id_idx] or "").strip()
        if row_question_id != question_id.strip():
            continue
        
        response_id = str(row[response_id_idx] or "").strip()
        response_text = str(row[response_text_idx] or "").strip()
        
        if not response_id:
            continue
        
        response_data = {
            "question_id": row_question_id,
            "response_id": response_id,
            "response_text": response_text,
        }
        
        # Collect all theme_ids from all model columns for this response
        theme_counts: Counter[str] = Counter()
        for col_idx in theme_col_indices:
            theme_id = str(row[col_idx] or "").strip()
            if theme_id:
                theme_counts[theme_id] += 1
        
        # Convert to sorted list of tuples (theme_id, count) sorted by count descending
        theme_assignments = sorted(theme_counts.items(), key=lambda x: x[1], reverse=True)
        
        results.append((response_data, theme_assignments))
    
    wb.close()
    return results


def load_theme_labels(workbook_path: str, question_id: str) -> Dict[str, str]:
    """Load theme_id -> theme_label mapping from theme_catalog.
    
    Returns: Dict mapping theme_id -> theme_label
    """
    themes = load_theme_catalog(workbook_path, question_id)
    return {theme.theme_id: theme.theme_label for theme in themes}


def create_review_assignments_sheet(
    workbook: str,
    question_id: str,
    verbose: int = 0,
) -> str:
    """Create review_assignments sheet in long/tidy format.
    
    Creates a sheet with one row per response-theme combination, sorted by:
    - Primary: response_id (matching original order)
    - Secondary: model_count (descending)
    
    Returns: Path to updated workbook
    """
    # Collect theme assignments per response
    if verbose >= 1:
        print(f"Collecting theme assignments for question_id '{question_id}'...", file=sys.stderr)
    
    response_assignments = collect_theme_assignments_per_response(workbook, question_id)
    
    if not response_assignments:
        if verbose >= 1:
            print("No theme assignments found in responses_coded", file=sys.stderr)
        # Still create the sheet with just headers
        wb = openpyxl.load_workbook(workbook)
        if "review_assignments" in wb.sheetnames:
            wb.remove(wb["review_assignments"])
        sheet = wb.create_sheet("review_assignments")
        sheet.append(["question_id", "response_id", "response_text", "theme_id", "theme_label", "model_count"])
        wb.save(workbook)
        wb.close()
        return workbook
    
    # Load theme labels
    theme_labels = load_theme_labels(workbook, question_id)
    
    if verbose >= 1:
        total_responses = len(response_assignments)
        total_assignments = sum(len(themes) for _, themes in response_assignments)
        print(f"Found {total_assignments} theme assignments across {total_responses} responses", file=sys.stderr)
    
    # Build rows for the sheet, tracking original response index for sorting
    rows: List[Tuple[int, List[Any]]] = []  # (original_index, row_data)
    
    for response_idx, (response_data, theme_assignments) in enumerate(response_assignments):
        question_id_val = response_data["question_id"]
        response_id_val = response_data["response_id"]
        response_text_val = response_data["response_text"]
        
        # Create one row per theme assignment
        for theme_id, model_count in theme_assignments:
            theme_label = theme_labels.get(theme_id, "")  # Empty string if theme not in catalog
            
            rows.append((
                response_idx,  # Original response index to preserve order
                [
                    question_id_val,
                    response_id_val,
                    response_text_val,
                    theme_id,
                    theme_label,
                    model_count,
                ]
            ))
    
    # Sort rows: primary by original response order, secondary by model_count (descending)
    rows.sort(key=lambda r: (r[0], -r[1][5]))  # r[0] is original_index, r[1][5] is model_count
    
    # Extract just the row data (without the index)
    sorted_rows = [row_data for _, row_data in rows]
    
    # Write to workbook
    wb = openpyxl.load_workbook(workbook)
    
    # Delete existing sheet if present
    if "review_assignments" in wb.sheetnames:
        wb.remove(wb["review_assignments"])
    
    sheet = wb.create_sheet("review_assignments")
    
    # Write header
    sheet.append(["question_id", "response_id", "response_text", "theme_id", "theme_label", "model_count"])
    
    # Write data rows
    for row in sorted_rows:
        sheet.append(row)
    
    wb.save(workbook)
    wb.close()
    
    if verbose >= 1:
        print(f"Created review_assignments sheet with {len(sorted_rows)} rows", file=sys.stderr)
    
    return workbook

